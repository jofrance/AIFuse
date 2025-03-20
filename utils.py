import os
import config
import processing
import csv
import pandas as pd
from openpyxl.styles import Alignment

def safe_read_csv(file_name):
    rows = []
    max_cols = 0
    with open(file_name, 'r', newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)
            if len(row) > max_cols:
                max_cols = len(row)
    if rows and rows[0]:
        header = rows[0]
        if len(header) < max_cols:
            header += [f"col{i}" for i in range(len(header), max_cols)]
        data = rows[1:]
    else:
        header = [f"col{i}" for i in range(max_cols)]
        data = rows
    padded_data = [row + [None]*(max_cols - len(row)) for row in data]
    df = pd.DataFrame(padded_data, columns=header)
    return df

def write_csv_to_excel(csv_file, excel_file):
    try:
        df = safe_read_csv(csv_file)
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Results')
            workbook = writer.book
            sheet = writer.sheets['Results']
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
                for cell in row:
                    cell.number_format = '@'
                    cell.alignment = Alignment(horizontal='left')
        print(f"Excel file written to {excel_file}")
    except Exception as e:
        print(f"Error writing Excel file: {e}")

def check_resume_status():
    """
    Checks the resume status by comparing the total input cases with the processed cases.
    If the processed tracking file does not exist or is empty, it returns that resume is not possible.
    
    Returns a dictionary with:
      - total_input: total number of cases from the input file.
      - processed_count: number of cases that have already been processed.
      - resume_possible: True if there are some (but not all) cases processed.
    """
    import os
    import config
    import processing

    # Count total input cases from the input file.
    try:
        with open(config.ARGS.file, 'r') as f:
            total_input = sum(1 for line in f if line.strip())
    except Exception:
        total_input = 0

    # If the tracking file doesn't exist, or exists but is empty, there's nothing to resume.
    if not os.path.exists(job.processed_tracking_file):
        return {"total_input": total_input, "processed_count": 0, "resume_possible": False}
    else:
        with open(job.processed_tracking_file, 'r') as f:
            processed_lines = [line.strip() for line in f if line.strip()]
        processed_count = len(processed_lines)
        # Only allow resume if at least one case has been processed and not all.
        resume_possible = total_input > 0 and processed_count > 0 and processed_count < total_input
        return {
            "total_input": total_input,
            "processed_count": processed_count,
            "resume_possible": resume_possible
        }

